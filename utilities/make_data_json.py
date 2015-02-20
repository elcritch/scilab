#!/usr/bin/env python3


import shutil, re, sys, os, itertools, argparse, json
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

import scilab.tools.scriptrunner as ScriptRunner
from scilab.tools.scriptrunner import RESEARCH, RAWDATA, debug

import logging

from scilab.tools.excel import *
from scilab.tools.project import *

import scilab.tools.json as Json
# from scilab.expers.mechanical.fatigue.uts import TestInfo
from scilab.expers.mechanical.fatigue.helpers import flatten
from scilab.tools.tables import mdBlock, mdHeader, ImageTable, MarkdownTable
    
## Main
    
def parse_fatigue_data_sheet_v1(ws):
    
    rng = rangerForRow(ws)

    data = {}
    data['info'] = DataTree()
    data['info'].update( dictFrom(rng('A1:B1')) )
    
    measurements = DataTree(withProperties='width depth area')
    measurements.width.value = ws['B6'].value
    measurements.depth.value = ws['C6'].value
    measurements.area.value = ws['E6'].value
    
    measurements.width.stdev = ws['B7'].value
    measurements.depth.stdev = ws['C7'].value
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    
    # Process all the values in these rows
    other = DataTree()
    
    ## continue reading the column down 
    end = process_definitions_column(ws, other, 'A',8,30, stop_key='Failure Notes / Test Results', dbg=False)
    ## read next definition column (4 excel columns over)
    process_definitions_column(ws, other, 'D', 7, end, dbg=True)

    debug(other)
    
    # gauge
    data['gauge'] = gauge = DataTree()
    gauge.units = 'mm' # default
    
    if 'gauge' in other:
        gauge.value = other.pop('gauge')
    
    if 'init_position' in other:
        gauge.init_position = other.pop('init_position')
    elif 'gauge_init' in other:
        gauge.init_position = other.pop('gauge_init')
    
    if 'gauge_base' in other:
        gauge.base = other.pop('gauge_base')
    elif 'base_position' in other:
        gauge.base = other.pop('base_position')
    else:
        logging.warn("Excel file missing gauge_base! Possible keys:\t"+str([ str(k) for k in other.keys() ]) )
    
    
    data['other'] = other
    
    debug(data.keys())
    # data.measurements.area.value = other.pop('area')
    
    notes = {}
    process_definitions_column(ws, notes, 'A', end+1,end+5, stop_key=None, dbg=None)
    data['notes'] = notes
    
    return data
    
def parse_data_from_worksheet(testinfo, testfolder, args, **kwargs):

    try:
        excelfile = testfolder.datasheet 
        excelfile = excelfile.resolve()
        debug(excelfile)
        
        wb = load_workbook(excelfile.absolute().as_posix(), data_only=True)
    except (Exception) as err:
        logging.warn("Cannot open file:\n\t"+str(excelfile)+"\n\t vs \n\t"+excelfile.absolute().as_posix())
        raise err
        
    ## Process Excel Sheets
    ws = wb.worksheets[0]
    return parse_fatigue_data_sheet_v1(ws)    

def process_image_measurements_v1(testinfo, imgdata):
    """ UTS based v1 """
    data = {}
    data['info'] = DataTree()
    data['info'].update( testinfo.as_dict() )
    data['info']['set'] = testinfo.name
    
    # debug('\n'.join(map(str,flatten(imgdata,sep='.').items())))
    
    measurements = DataTree(width=DataTree(), depth=DataTree(), area=DataTree())
    measurements.width.value = imgdata.front.widths.average.mean
    
    if not 'side' in imgdata:
        logging.warn("Could not find side measurement for: "+str(testinfo))
        measurements.depth.value = 1.00
        measurements.depth.stdev = -1.0
    else:
        measurements.depth.value = imgdata.side.widths.average.mean
        measurements.depth.stdev = imgdata.side.widths.average.std
        
    measurements.area.value = float(measurements.width.value)*float(measurements.depth.value)
    
    measurements.width.stdev = imgdata.front.widths.average.std
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    measurements.image = DataTree(other={})
    
    data['measurements'] = measurements
    
    return data

def process_image_measurements(testinfo, imgdata):
        
    data = {}
    data['info'] = DataTree()
    data['info'].update( testinfo.as_dict() )
    data['info']['set'] = testinfo.name
    
    # debug('\n'.join(map(str,flatten(imgdata,sep='.').items())))
    
    measurements = DataTree(width=DataTree(), depth=DataTree(), area=DataTree())
    measurements.width.value = imgdata.front.mm.summaries.widths.average.mean
    
    if not 'side' in imgdata:
        logging.warn("Could not find side measurement for: "+str(testinfo))
        measurements.depth.value = 1.00
        measurements.depth.stdev = -1.0
    else:
        measurements.depth.value = imgdata.side.mm.summaries.widths.average.mean
        measurements.depth.stdev = imgdata.side.mm.summaries.widths.average.std
        
    measurements.area.value = float(measurements.width.value)*float(measurements.depth.value)
    
    measurements.width.stdev = imgdata.front.mm.summaries.widths.average.std
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    measurements.image = DataTree(other={})
    
    for k in 'front side fail'.split():
        measurements.image.other[k] = imgdata[k].mm.other if k in imgdata else {}
    
    data['measurements'] = measurements
    
    return data


def parse_from_image_measurements(testinfo, testfolder, args):

    # imgMeasureFile = testfolder.image / 'processed' / (testinfo.name + '.measurements.json')
    imgMeasureFile = testfolder.images / 'processed' / 'data.json'
    imgMeasureFile = imgMeasureFile.resolve()        
    
    imgMeasurements = Json.load_json(imgMeasureFile.parent.as_posix(), json_url=imgMeasureFile.name)
    # debug(imgMeasurements)
    
    try:
        data = process_image_measurements(testinfo, imgMeasurements)
    except:
        data = process_image_measurements_v1(testinfo, imgMeasurements)

    return data

def graphs2_handler(testinfo, testfolder, args, testdata, **kwargs):
    
    excelfile = testdata.datasheet
    
    debug(excelfile)
    try:
        handler(testinfo, testfolder, excelfile=excelfile, args=args, )
    except FileNotFoundError as err:
        logging.error("FileNotFoundError:", testinfo.name, err)
    
    

def handler(testinfo, testfolder, excelfile, args):
    
    
    print(mdHeader(2, "Test: "+testinfo.name), file=args.report)

    def updateMetaData(data):
        ## Handle Names    
        data['name'] = testinfo.name        
        data['id'] = testinfo.short()
        data['info'] = testinfo.as_dict()
    
    
    debug(testfolder.jsoncalc.as_posix())
    
    print(str(testinfo), file=args.report)
    
    ## Update with Excel Values    
    data = parse_data_from_worksheet(
        testpath=excelfile,
        testinfo=testinfo,
        testfolder=testfolder,
        args=args)

    print(mdBlock("Excel Sheet Data:"),file=args.report)
    print(mdBlock("```json\n"+json.dumps(data,indent=4)+"\n```"),file=args.report)
    updateMetaData(data)


    Json.write_json(
        testfolder.jsoncalc.as_posix(),
        data,
        json_url=testinfo.name+'.excel.calculated.json',
        dbg=False)

    
    ## Update with Image Measurements
    
    data = parse_from_image_measurements(
        testinfo=testinfo,
        testfolder=testfolder,
        args=args)
    
    updateMetaData(data)

    Json.write_json(
        testfolder.jsoncalc.as_posix(),
        data,
        json_url=testinfo.name+'.measurements.calculated.json',
        dbg=False)
    
    return

if __name__ == '__main__':

    parser = ScriptRunner.parser
    
    parser.add_argument("-n", "--name", default="data", type=str)

    ## Test

    projectspath = Path(RESEARCH) / '07_Experiments'
    projectpath = projectspath/'fatigue failure (UTS, exper1)'
    
    experData = projectpath / 'test-data'/'uts (expr-1)'
    experExcel = experData/'01 Excel' 
    experJson = experData/'00 JSON'
    experReport = experData/'02 Reports'
    
    files = experExcel.glob('*.xlsx')

    test_args = []
    # test_args = ["--glob", fileglob]
    # test_args += ['-1'] # only first
    
    args = parser.parse_args( test_args )    
    args = parser.parse_args( test_args )    

    experUtsCsv = projectpath / '04 (uts) uts-test' 
    experUtsPreconds = projectpath / '02 (uts) preconditions' 
    experData = projectpath / 'test-data'/'uts (expr-1)'
    experExcel = experData/'01 Excel' 
    experJson = experData/'00 JSON'
    experReport = experData/'02 Reports'
    experReportGraphs = experData/'03 Graphs'
    experJsonCalc = experJson / 'calculated'
    [ setattr(args,e,v) for e,v in locals().items() if e.startswith('exper' )]
    
    files = experExcel.glob('*.xlsx')    
    
    with (args.experReport/'Temp Reports'/'Excel Data Sheet Results.md').open('w') as report:
        
        class tee:
            def __init__(self, _fd1, _fd2) :
                self.fd1 = _fd1
                self.fd2 = _fd2

            def __del__(self) :
                if self.fd1 != sys.stdout and self.fd1 != sys.stderr :
                    self.fd1.close()
                if self.fd2 != sys.stdout and self.fd2 != sys.stderr :
                    self.fd2.close()

            def write(self, text) :
                self.fd1.write(text)
                self.fd2.write(text)

            def flush(self) :
                self.fd1.flush()
                self.fd2.flush()

        args.report = tee(sys.stdout, report)

        
        for test in list(files)[:]:
        
            debug(test)
        
            handler(file=test, args=args)
    
    
    
