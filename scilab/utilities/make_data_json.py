#!/usr/bin/env python3


import shutil, re, sys, os, itertools, argparse, json, collections
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

import scilab.tools.scriptrunner as ScriptRunner
# from scilab.tools.scriptrunner import RESEARCH, RAWDATA, debug

import logging

from scilab.tools.excel import *
from scilab.tools.project import *

import scilab.tools.jsonutils as Json
# from scilab.expers.mechanical.fatigue.uts import TestInfo
# from scilab.expers.mechanical.fatigue.helpers import flatten
from scilab.tools.tables import mdBlock, mdHeader, ImageTable, MarkdownTable
    
## Main
    
def parse_fatigue_data_sheet_v1(ws):
    
    rng = rangerForRow(ws)

    data = DataTree()
    data['info'] = DataTree()
    data['info'].update( dictFrom(rng('A1:B1')) )
    
    measurements = DataTree(withProperties='width depth area')
    measurements.width.value = ws['B6'].value
    measurements.depth.value = ws['C6'].value
    measurements.area.value = ws['E6'].value
    
    measurements.width.stdev = ws['B7'].value
    measurements.depth.stdev = ws['C7'].value
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    
    # Process all the values in these rows
    other = DataTree()
    
    ## continue reading the column down 
    end = process_definitions_column(ws, other, 'A',8,50, stop_key='Failure Notes / Test Results', dbg=False)
    ## read next definition column (4 excel columns over)
    process_definitions_column(ws, other, 'D', 7, end, dbg=True)

    debug(other)
    
    # gauge
    gauge = DataTree()
    units = 'mm' # default
    
    if 'gauge' in other:
        gauge.length = valueUnits(other.pop('gauge', units), units)._asdict()
    
    if 'init_position' in other:
        gauge.init_position = valueUnits(other.pop('init_position'), units)._asdict()
    elif 'gauge_init' in other:
        gauge.init_position = valueUnits(other.pop('gauge_init'), units)._asdict()
    
    if 'gauge_base' in other:
        gauge.base = valueUnits(other.pop('gauge_base'), units)._asdict()
    elif 'base_position' in other:
        gauge.base = valueUnits(other.pop('base_position'), units)._asdict()
    else:
        logging.warn("Excel file missing gauge_base! Possible keys:\t"+str([ str(k) for k in other.keys() ]) )
    
    data['excel','other'] = other
    data['measurements'] = DataTree(gauge=gauge)
    
    debug(data.keys())
    # data.measurements.area.value = other.pop('area')
    
    notes = {}
    process_definitions_column(ws, notes, 'A', end+1,end+5, stop_key=None, dbg=None)
    data['excel','notes'] = notes
    
    return data
    
def parse_data_from_worksheet(testconf, args, **kwargs):

    try:
        excelfile = testconf.folder.datasheet 
        excelfile = excelfile.resolve()
        debug(excelfile)
        
        wb = load_workbook(excelfile.absolute().as_posix(), data_only=True)
    except (Exception) as err:
        logging.warn("Cannot open file:\n\t"+str(excelfile)+"\n\t vs \n\t"+str(excelfile))
        return 
        
    ## Process Excel Sheets
    ws = wb.worksheets[0]
    return parse_fatigue_data_sheet_v1(ws)    

def process_image_measurements_v1(testconf, imgdata):
    """ UTS based v1 """
    data = {}
    data['info'] = DataTree()
    data['info'].update( testconf.info.as_dict() )
    data['info']['set'] = testconf.info.name
    
    # debug('\n'.join(map(str,flatten(imgdata,sep='.').items())))
    
    measurements = DataTree(width=DataTree(), depth=DataTree(), area=DataTree())
    measurements.width.value = imgdata.front.widths.average.mean
    
    if not 'side' in imgdata:
        logging.error("Could not find side measurement for: "+str(testconf.info))
        measurements.depth.value = 1.00
        measurements.depth.stdev = -1.0
    else:
        measurements.depth.value = imgdata.side.widths.average.mean
        measurements.depth.stdev = imgdata.side.widths.average.std
        
    measurements.area.value = float(measurements.width.value)*float(measurements.depth.value)
    
    measurements.width.stdev = imgdata.front.widths.average.std
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    measurements.image = DataTree(other={})
    
    data['measurements'] = measurements
    
    return data

def process_image_measurements(testconf, imgdata):
        
    data = {}
    data['info'] = DataTree()
    data['info'].update( testconf.info.as_dict() )
    data['info']['set'] = testconf.info.name
    
    # debug('\n'.join(map(str,flatten(imgdata,sep='.').items())))
    
    measurements = DataTree(width=DataTree(), depth=DataTree(), area=DataTree())
    measurements.width.value = imgdata.front.mm.summaries.widths.average.mean
    
    if not 'side' in imgdata:
        logging.warn("Could not find side measurement for: "+str(testconf.info))
        measurements.depth.value = 1.00
        measurements.depth.stdev = -1.0
    else:
        measurements.depth.value = imgdata.side.mm.summaries.widths.average.mean
        measurements.depth.stdev = imgdata.side.mm.summaries.widths.average.std
        
    measurements.area.value = float(measurements.width.value)*float(measurements.depth.value)
    
    measurements.width.stdev = imgdata.front.mm.summaries.widths.average.std
    
    measurements.width.units = 'mm'
    measurements.depth.units = 'mm'
    measurements.area.units = 'mm^2'
    
    measurements.image = DataTree(other={})
    
    for k in 'front side fail'.split():
        measurements.image.other[k] = imgdata[k].mm.other if k in imgdata else {}
    
    data['measurements'] = measurements
    
    return data


def parse_from_image_measurements(testconf, args):

    # imgMeasureFile = testconf.folder.image / 'processed' / (testconf.info.name + '.measurements.json')
    imgMeasureFile = testconf.folder.images / 'processed' / 'data.json'
    imgMeasureFile = imgMeasureFile.resolve()        
    
    imgMeasurements = Json.load_json(imgMeasureFile.parent.as_posix(), json_url=imgMeasureFile.name)
    # debug(imgMeasurements)
    
    try:
        data = process_image_measurements(testconf, imgMeasurements)
    except:
        data = process_image_measurements_v1(testconf, imgMeasurements)

    return data

def graphs2_handler(testconf, args, **kwargs):
    
    excelfile = testconf.folder.datasheet
    
    debug(excelfile)
    try:
        handler(testconf.info, testconf.folder, excelfile=excelfile, args=args, )
    except FileNotFoundError as err:
        logging.error("FileNotFoundError:", testconf.info.name, err)
    except:
        return
    
def HTML(arg):
    return arg.replace('\n','')

def handler(testconf, excelfile, args):
    
    
    # print(mdHeader(3, ": {}",testconf.info.name))

    def updateMetaData(data):
        ## Handle Names    
        data['name'] = testconf.info.name
        data['id'] = testconf.info.short
        data['info'] = testconf.info.as_dict()
    
    
    debug(testconf.folder.jsoncalc.as_posix())
    
    # print(str(testconf.info), file=args.report)
    
    ## Update with Excel Values    
    data = parse_data_from_worksheet(
        testpath=excelfile,
        testconf=testconf,
        args=args,
        )

    print(mdBlock("Excel Sheet Data:"))
    rows = [ (k,v) for k,v in flatten(data).items() ]
    rows = sorted(rows)

    print()
    print(HTML(tabulate.tabulate( rows, [ "Key", "Value", ], tablefmt ='html' )))
    
    # print(mdBlock("<pre>\n"+json.dumps(data,indent=4)+"\n</pre>"))
    updateMetaData(data)
    
    # excel_data = DataTree(notes=data.pop("notes"), other=data.pop("other"))
    json_data = collections.OrderedDict(sorted(data.items()))
    testconf.folder.save_calculated_json_raw(test=testconf, name='excel',json_data=json_data, overwrite=True)
    
    ## Update with Image Measurements
    
    data = parse_from_image_measurements(
        testconf=testconf,
        args=args)
    
    updateMetaData(data)

    testconf.folder.save_calculated_json_raw(test=testconf, name='measurements',json_data=data, overwrite=True)
    
    return

if __name__ == '__main__':

    parser = ScriptRunner.parser
    
    parser.add_argument("-n", "--name", default="data", type=str)

    ## Test

    projectspath = Path(RESEARCH) / '07_Experiments'
    projectpath = projectspath/'fatigue failure (UTS, exper1)'
    
    experData = projectpath / 'test-data'/'uts (expr-1)'
    experExcel = experData/'01 Excel' 
    experJson = experData/'00 JSON'
    experReport = experData/'02 Reports'
    
    files = experExcel.glob('*.xlsx')

    testconf._args = []
    # testconf._args = ["--glob", fileglob]
    # testconf._args += ['-1'] # only first
    
    args = parser.parse_args( testconf._args )    
    args = parser.parse_args( testconf._args )    

    experUtsCsv = projectpath / '04 (uts) uts-test' 
    experUtsPreconds = projectpath / '02 (uts) preconditions' 
    experData = projectpath / 'test-data'/'uts (expr-1)'
    experExcel = experData/'01 Excel' 
    experJson = experData/'00 JSON'
    experReport = experData/'02 Reports'
    experReportGraphs = experData/'03 Graphs'
    experJsonCalc = experJson / 'calculated'
    [ setattr(args,e,v) for e,v in locals().items() if e.startswith('exper' )]
    
    files = experExcel.glob('*.xlsx')    
    
    with (args.experReport/'Temp Reports'/'Excel Data Sheet Results.md').open('w') as report:
        
        class tee:
            def __init__(self, _fd1, _fd2) :
                self.fd1 = _fd1
                self.fd2 = _fd2

            def __del__(self) :
                if self.fd1 != sys.stdout and self.fd1 != sys.stderr :
                    self.fd1.close()
                if self.fd2 != sys.stdout and self.fd2 != sys.stderr :
                    self.fd2.close()

            def write(self, text) :
                self.fd1.write(text)
                self.fd2.write(text)

            def flush(self) :
                self.fd1.flush()
                self.fd2.flush()

        args.report = tee(sys.stdout, report)

        
        for test in list(files)[:]:
        
            debug(test)
        
            handler(file=test, args=args)
    
    
    
